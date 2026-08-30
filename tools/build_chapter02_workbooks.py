"""Build the Chapter 2 demonstration workbooks in data/.

Three files, all built from data/state_hpi_ur_pop.csv:

    ch02-frequency-BROKEN.xlsx    what a loose prompt actually returns
    ch02-frequency-FIXED.xlsx     what a specified prompt returns
    ch02-frequency-STARTER.xlsx   the data and the checks, nothing filled in

The point of the pair is not that AI is unreliable. It is that a frequency
distribution has rules -- mutually exclusive classes, all n values accounted
for, readable class boundaries -- and those rules are the specification you
hand the assistant AND the checklist you run on what comes back. The BROKEN
workbook breaks three of them at once, in the three ways this task actually
fails in practice:

  1. The frequency, relative and cumulative columns are PASTED NUMBERS.
     Change a class boundary and nothing moves. There is no way to tell by
     looking; you have to click a cell.

  2. The class boundaries were derived straight from (max - min) / k and
     never rounded, so they read -4.20 to 0.30 instead of -5 to less than 0.
     Donnelly p.34: round the estimate to a useful whole number.

  3. The boundaries are inclusive at BOTH ends, so a state sitting exactly on
     a boundary lands in two classes. South Carolina grew 13.8%, which is
     exactly the fourth boundary, so it is counted twice and the frequencies
     sum to 51 instead of 50. Donnelly's rule 2: mutually exclusive classes.

Every workbook carries a Checks sheet holding the six verification formulas
from the lecture note. They are live formulas, not commentary: open BROKEN and
the sheet is already reporting FAIL before you have looked at anything else.

Run from the repository root:

    python tools/build_chapter02_workbooks.py

Requires openpyxl (already in requirements.txt).

NOTE ON CACHED VALUES: openpyxl writes formulas, not results. Excel, LibreOffice
and Google Sheets all evaluate them on open. A plain-text viewer will show the
formulas themselves, which for this exercise is arguably the honest display.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
SOURCE = REPO / "data" / "state_hpi_ur_pop.csv"
OUT = REPO / "data"

# The column of state_hpi_ur_pop.csv we build the distribution from.
VAR = "gPOP"
VAR_LABEL = "Population growth over the decade (%)"

# The readable class scheme: 2^6 = 64 >= 50, so k = 6; range is -4.2 to 22.8,
# so (22.8 - -4.2) / 6 = 4.5, rounded up to a readable width of 5 starting at
# a readable -5. This is the textbook's procedure, done properly.
FIXED_BOUNDS = [-5, 0, 5, 10, 15, 20, 25]

SCARLET = "A03123"
GRAY = "F2F2F2"

HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor=SCARLET)
NOTE = Font(italic=True, color="666666", size=9)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BBBBBB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------- helpers

def _header(ws, row, labels, start=1):
    for j, label in enumerate(labels):
        c = ws.cell(row=row, column=start + j, value=label)
        c.font, c.fill, c.border = HEAD, HEAD_FILL, BOX
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def _title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=13, color=SCARLET)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = NOTE


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def data_sheet(wb, df):
    """The raw file, headers in row 1, nothing else. Same in all three books."""
    ws = wb.create_sheet("Data")
    ws.append(list(df.columns))
    for c in ws[1]:
        c.font, c.fill = HEAD, HEAD_FILL
    for row in df.itertuples(index=False):
        ws.append([str(v) if isinstance(v, str) else v for v in row])
    _widths(ws, [9] + [10] * (len(df.columns) - 1))
    ws.freeze_panes = "A2"
    return ws


# ------------------------------------------------------------ checks sheet

# Every check is a live formula, so it keeps checking after you edit the
# workbook. Checks 5 and 6 are meaningless on an empty table, so they report
# "not yet" rather than a false PASS -- a checklist that passes a blank sheet
# is worse than no checklist.
_EMPTY = 'IF(N({freq_total})=0,"-- not yet --",{body})'

CHECKS = [
    ("Frequencies sum to n",
     "Donnelly rule 3: include all data values",
     '=IF(N({freq_total})={n},"PASS","FAIL: "&TEXT(N({freq_total}),"0")'
     '&" of {n} accounted for")'),
    ("Relative frequencies sum to 1.00",
     "p.28",
     '=IF(ABS(N({rel_total})-1)<0.005,"PASS","FAIL: sums to "'
     '&TEXT(N({rel_total}),"0.000"))'),
    ("Cumulative column ends at 1.00",
     "p.28",
     '=IF(ABS(N({cum_last})-1)<0.005,"PASS","FAIL: ends at "'
     '&TEXT(N({cum_last}),"0.000"))'),
    ("Frequency column holds formulas, not typed numbers",
     "the reproducibility rule: change a bound, watch it move",
     '=IF(_xlfn.ISFORMULA({freq_first}),"PASS",IF(ISBLANK({freq_first}),'
     '"-- not yet --","FAIL: {freq_first} is a typed number, not a formula"))'),
    ("Classes are mutually exclusive (no value in two classes)",
     "Donnelly rule 2",
     "=" + _EMPTY.format(
         freq_total="{freq_total}",
         body='IF(N({freq_total})<={n},"PASS",'
              '"FAIL: "&TEXT(N({freq_total})-{n},"0")&" value(s) counted twice")')),
    ("No empty classes",
     "Donnelly rule 4",
     "=" + _EMPTY.format(
         freq_total="{freq_total}",
         body='IF(COUNTIF({freq_range},0)=0,"PASS","FAIL: "&'
              'TEXT(COUNTIF({freq_range},0),"0")&" empty class(es)")')),
]


def checks_sheet(wb, refs, *, blurb):
    ws = wb.create_sheet("Checks")
    _title(ws, "Verification checklist",
           "Six checks, each under a minute. Every one is a rule from Chapter 2, "
           "written as a live formula. Do not accept a workbook until all six read PASS.")
    _header(ws, 4, ["#", "Check", "What rule it enforces", "Result"])
    for i, (name, rule, formula) in enumerate(CHECKS, start=1):
        r = 4 + i
        ws.cell(row=r, column=1, value=i).border = BOX
        ws.cell(row=r, column=2, value=name).border = BOX
        c = ws.cell(row=r, column=3, value=rule)
        c.font, c.border = NOTE, BOX
        v = ws.cell(row=r, column=4, value=formula.format(**refs))
        v.border, v.font = BOX, BOLD
    ws.cell(row=12, column=1, value=blurb).font = NOTE
    ws.cell(row=12, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=12, start_column=1, end_row=16, end_column=4)
    _widths(ws, [4, 46, 44, 46])
    return ws


# ------------------------------------------------------------------ BROKEN

def build_broken(df):
    """What you get from 'make me a frequency distribution of this in Excel'."""
    g = df[VAR]
    lo, hi = float(g.min()), float(g.max())
    width = (hi - lo) / 6
    bounds = [round(lo + i * width, 4) for i in range(7)]

    wb = Workbook()
    wb.remove(wb.active)
    data_sheet(wb, df)

    ws = wb.create_sheet("Distribution", 1)
    _title(ws, "Frequency distribution of state population growth",
           "Produced from a one-line request. Every number below is correct "
           "arithmetic on the wrong specification.")
    _header(ws, 4, ["Class", "Lower", "Upper", "Frequency",
                    "Relative frequency", "Cumulative relative frequency"])

    cum = 0.0
    for i in range(6):
        r = 5 + i
        # Boundaries inclusive at BOTH ends -- this is the double count.
        n = int(((g >= bounds[i]) & (g <= bounds[i + 1])).sum())
        cum += n / len(g)
        vals = [f"{bounds[i]:.2f} to {bounds[i+1]:.2f}", bounds[i], bounds[i + 1],
                n,                       # pasted, not =COUNTIFS(...)
                round(n / len(g), 4),    # pasted
                round(cum, 4)]           # pasted
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BOX
            if j in (5, 6):
                c.number_format = "0.000"

    total_row = 11
    ws.cell(row=total_row, column=1, value="Total").font = BOLD
    # Even the total is pasted, so it agrees with itself and with nothing else.
    ws.cell(row=total_row, column=4,
            value=int(sum(((g >= bounds[i]) & (g <= bounds[i + 1])).sum()
                          for i in range(6)))).font = BOLD
    ws.cell(row=total_row, column=5, value=round(cum, 4)).font = BOLD
    ws.cell(row=total_row, column=5).number_format = "0.000"
    for j in range(1, 7):
        ws.cell(row=total_row, column=j).border = BOX

    ws.cell(row=13, column=1,
            value="There are 50 states. Add up the Frequency column.")
    ws.cell(row=13, column=1).font = Font(italic=True, color=SCARLET, size=10)

    _widths(ws, [20, 10, 10, 12, 20, 30])
    _bar(ws, "Population growth, six classes", rows=(5, 10))

    checks_sheet(wb, dict(
        n=len(g), freq_total=f"Distribution!D{total_row}",
        rel_total=f"SUM(Distribution!E5:E10)",
        cum_last="Distribution!F10", freq_first="Distribution!D5",
        freq_range="Distribution!D5:D10",
    ), blurb=(
        "Three things went wrong here, and a one-line prompt is why.\n\n"
        "1. Nothing is a formula. Click D5. Change a boundary in column B and "
        "watch nothing happen.\n"
        "2. The class boundaries came straight from (max - min) / 6 and were "
        "never rounded, so they read -4.20 to 0.30 instead of -5 to less than 0. "
        "Donnelly p.34 says round the estimate to a readable whole number.\n"
        "3. The boundaries include BOTH endpoints, so South Carolina -- which "
        "grew exactly 13.8%, exactly a boundary -- is counted in two classes. "
        "Fifty states, fifty-one observations. That is Donnelly's rule 2.\n\n"
        "None of this is visible in the chart, and the chart is the only part "
        "most people look at."))
    return wb


# ------------------------------------------------------------------- FIXED

def build_fixed(df):
    """What you get when the prompt names every decision the textbook makes."""
    n = len(df)
    col = get_column_letter(list(df.columns).index(VAR) + 1)   # gPOP in Data
    rng = f"Data!${col}$2:${col}${n + 1}"

    wb = Workbook()
    wb.remove(wb.active)
    data_sheet(wb, df)

    ws = wb.create_sheet("Distribution", 1)
    _title(ws, "Frequency distribution of state population growth",
           "Six classes, width 5, half-open boundaries. Every count is a live "
           "COUNTIFS against the Data sheet. Edit a bound in column B or C and "
           "the table, the total and both charts follow.")
    _header(ws, 4, ["Class", "Lower", "Upper (less than)", "Frequency",
                    "Relative frequency", "Cumulative relative frequency"])

    first, last = 5, 10
    for i in range(6):
        r = first + i
        lo, hi = FIXED_BOUNDS[i], FIXED_BOUNDS[i + 1]
        ws.cell(row=r, column=1,
                value=f'=B{r}&" to less than "&C{r}').border = BOX
        ws.cell(row=r, column=2, value=lo).border = BOX
        ws.cell(row=r, column=3, value=hi).border = BOX
        # Half-open: >= lower, < upper. No value can land in two classes.
        ws.cell(row=r, column=4,
                value=f'=COUNTIFS({rng},">="&B{r},{rng},"<"&C{r})').border = BOX
        e = ws.cell(row=r, column=5, value=f"=D{r}/$D${last + 1}")
        e.border, e.number_format = BOX, "0.000"
        f = ws.cell(row=r, column=6,
                    value=(f"=E{r}" if i == 0 else f"=F{r - 1}+E{r}"))
        f.border, f.number_format = BOX, "0.000"

    total_row = last + 1
    ws.cell(row=total_row, column=1, value="Total").font = BOLD
    t = ws.cell(row=total_row, column=4, value=f"=SUM(D{first}:D{last})")
    t.font = BOLD
    tr = ws.cell(row=total_row, column=5, value=f"=SUM(E{first}:E{last})")
    tr.font, tr.number_format = BOLD, "0.000"
    for j in range(1, 7):
        ws.cell(row=total_row, column=j).border = BOX

    ws.cell(row=total_row + 2, column=1,
            value=(f"n in the source file: =COUNT({rng})  ->"))
    ws.cell(row=total_row + 2, column=1).font = NOTE
    ws.cell(row=total_row + 2, column=4, value=f"=COUNT({rng})").font = BOLD

    ws.cell(row=total_row + 3, column=1,
            value="Independent re-derivation of class 1, a different way:")
    ws.cell(row=total_row + 3, column=1).font = NOTE
    ws.cell(row=total_row + 3, column=4,
            value=f'=SUMPRODUCT(({rng}>=B{first})*({rng}<C{first}))').font = BOLD

    _widths(ws, [24, 10, 18, 12, 20, 30])
    _bar(ws, "Distribution of state population growth", rows=(first, last),
         gap=0)
    _ogive(ws, rows=(first, last))

    _units_sheet(wb, df)

    checks_sheet(wb, dict(
        n=n, freq_total=f"Distribution!D{total_row}",
        rel_total=f"Distribution!E{total_row}",
        cum_last=f"Distribution!F{last}", freq_first=f"Distribution!D{first}",
        freq_range=f"Distribution!D{first}:D{last}",
    ), blurb=(
        "All six should read PASS. Now make them fail on purpose.\n\n"
        "Change ONE cell on the Distribution sheet: C10, the top boundary, "
        "from 25 to 22. Idaho grew 22.8%, so it now falls outside every class. "
        "Without touching anything else, the total drops to 49, the last class "
        "empties, both charts redraw, and checks 1 and 6 turn red.\n\n"
        "Now look at checks 2 and 3. They still read PASS. The relative "
        "frequencies are computed against the total of the classes, not "
        "against n, so they renormalise to the 49 states you kept and sum to "
        "exactly 1.000. Every percentage on that sheet is now internally "
        "consistent and describes 49 states while claiming to describe 50.\n\n"
        "That is the Q38 pie chart from the lecture note, in a spreadsheet: "
        "correct arithmetic on a base nobody stated. It is also why check 1 "
        "compares against n from the source file rather than against the "
        "column's own total -- a checklist that only asks whether a sheet "
        "agrees with itself will pass this every time.\n\n"
        "Set C10 back to 25 before you go on."))
    return wb


def _units_sheet(wb, df):
    """The index-level trap, computed live rather than asserted."""
    n = len(df)
    cols = list(df.columns)
    L = {c: get_column_letter(cols.index(c) + 1) for c in
         ("HPI", "gHPI", "gPOP", "UR")}
    R = {k: f"Data!${v}$2:${v}${n + 1}" for k, v in L.items()}

    ws = wb.create_sheet("Units")
    _title(ws, "House price INDEX versus house price CHANGE",
           "The same fifty states, the same source, two different questions. "
           "Both correlations below are computed live from the Data sheet.")

    _header(ws, 4, ["Pairing", "Correlation", "What it is measuring"])
    rows = [
        ("Population growth vs house price INDEX LEVEL",
         f"=CORREL({R['gPOP']},{R['HPI']})",
         "The index is 1980Q1 = 100, so its level is cumulative appreciation "
         "since 1980 -- not the price of a house."),
        ("Population growth vs TEN-YEAR house price CHANGE",
         f"=CORREL({R['gPOP']},{R['gHPI']})",
         "Both sides are now a ten-year change. Same window, comparable units."),
        ("Unemployment rate vs ten-year house price change",
         f"=CORREL({R['UR']},{R['gHPI']})",
         "The null case. This is what no relationship looks like."),
    ]
    for i, (label, formula, why) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).border = BOX
        c = ws.cell(row=r, column=2, value=formula)
        c.border, c.font, c.number_format = BOX, BOLD, "+0.000;-0.000"
        w = ws.cell(row=r, column=3, value=why)
        w.border, w.font = BOX, NOTE
        w.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 32

    ws.cell(row=10, column=1, value=(
        "Nothing about the data changed between rows 5 and 6. The units did. "
        "Massachusetts tops the index at 1313 while Texas sits at 529 -- and "
        "Texas grew 15.4% while Massachusetts grew 5.2%. An index level and an "
        "index change are not interchangeable, and only one of them is "
        "comparable across states."))
    ws.cell(row=10, column=1).font = Font(italic=True, color=SCARLET, size=10)
    ws.cell(row=10, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=10, start_column=1, end_row=13, end_column=3)
    _widths(ws, [48, 14, 60])

    for i, (yc, title, anchor) in enumerate([
            ("HPI", "Population growth vs house price index LEVEL", "A16"),
            ("gHPI", "Population growth vs ten-year house price CHANGE", "J16")]):
        ch = ScatterChart()
        ch.title = title
        ch.x_axis.title = "Population growth over the decade (%)"
        ch.y_axis.title = "House price index" if yc == "HPI" else "Ten-year change (%)"
        ch.style = 13
        ch.height, ch.width = 8, 12
        xs = Reference(wb["Data"], min_col=cols.index("gPOP") + 1,
                       min_row=2, max_row=n + 1)
        ys = Reference(wb["Data"], min_col=cols.index(yc) + 1,
                       min_row=1, max_row=n + 1)
        s = Series(ys, xs, title_from_data=True)
        s.marker.symbol = "circle"
        s.graphicalProperties.line.noFill = True
        ch.series.append(s)
        ws.add_chart(ch, anchor)
    return ws


# ----------------------------------------------------------------- STARTER

def build_starter(df):
    """The data, the class scheme, the checks -- and an empty table."""
    n = len(df)
    wb = Workbook()
    wb.remove(wb.active)
    data_sheet(wb, df)

    ws = wb.create_sheet("Distribution", 1)
    _title(ws, "Frequency distribution of state population growth -- BUILD THIS",
           "The class boundaries are given. Columns D, E and F are yours. Use "
           "live formulas, not typed numbers: the Checks sheet can tell the "
           "difference.")
    _header(ws, 4, ["Class", "Lower", "Upper (less than)", "Frequency",
                    "Relative frequency", "Cumulative relative frequency"])
    for i in range(6):
        r = 5 + i
        ws.cell(row=r, column=1,
                value=f'=B{r}&" to less than "&C{r}').border = BOX
        ws.cell(row=r, column=2, value=FIXED_BOUNDS[i]).border = BOX
        ws.cell(row=r, column=3, value=FIXED_BOUNDS[i + 1]).border = BOX
        for j in (4, 5, 6):
            c = ws.cell(row=r, column=j)
            c.border = BOX
            c.fill = PatternFill("solid", fgColor=GRAY)
    ws.cell(row=11, column=1, value="Total").font = BOLD
    for j in range(1, 7):
        ws.cell(row=11, column=j).border = BOX
    ws.cell(row=13, column=1, value=(
        "Hint: the count for a half-open class is a COUNTIFS with two "
        "conditions -- at least the lower bound, and strictly less than the "
        "upper. Point it at the Data sheet, and reference the bound CELLS so "
        "you can change them later."))
    ws.cell(row=13, column=1).font = NOTE
    ws.cell(row=13, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=13, start_column=1, end_row=15, end_column=6)
    _widths(ws, [24, 10, 18, 12, 20, 30])

    checks_sheet(wb, dict(
        n=n, freq_total="Distribution!D11", rel_total="SUM(Distribution!E5:E10)",
        cum_last="Distribution!F10", freq_first="Distribution!D5",
        freq_range="Distribution!D5:D10",
    ), blurb=(
        "Every check reads FAIL right now, because the table is empty. That is "
        "the point: these six formulas tell you when you are finished, and they "
        "keep telling you every time you change something afterwards.\n\n"
        "Build the table on the Distribution sheet until all six read PASS."))
    return wb


# ------------------------------------------------------------------ charts

def _bar(ws, title, rows, gap=None, anchor="H4"):
    ch = BarChart()
    ch.type, ch.style = "col", 10
    ch.title = title
    ch.y_axis.title = "Frequency"
    ch.x_axis.title = VAR_LABEL
    if gap is not None:
        # Continuous data: no gaps between the bars. Donnelly p.37.
        ch.gapWidth = gap
    ch.add_data(Reference(ws, min_col=4, min_row=4, max_row=rows[1]),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=rows[0], max_row=rows[1]))
    ch.height, ch.width = 8.5, 15
    ws.add_chart(ch, anchor)


def _ogive(ws, rows, anchor="H22"):
    ch = LineChart()
    ch.title = "Ogive: cumulative relative frequency"
    ch.y_axis.title = "Cumulative relative frequency"
    ch.x_axis.title = VAR_LABEL
    ch.add_data(Reference(ws, min_col=6, min_row=4, max_row=rows[1]),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=rows[0], max_row=rows[1]))
    ch.height, ch.width = 8.5, 15
    ws.add_chart(ch, anchor)


# -------------------------------------------------------------------- main

def main():
    if not SOURCE.exists():
        raise SystemExit(f"Missing {SOURCE} -- run tools/seed_chapter02_data.R first")
    df = pd.read_csv(SOURCE)
    for col in ("HPI", "gHPI", "gPOP", "UR"):
        if col not in df.columns:
            raise SystemExit(
                f"{SOURCE.name} has no {col} column. Re-run "
                "tools/seed_chapter02_data.R -- the workbooks need the "
                "ten-year house price change as well as the index level.")

    builds = [("ch02-frequency-BROKEN.xlsx", build_broken),
              ("ch02-frequency-FIXED.xlsx", build_fixed),
              ("ch02-frequency-STARTER.xlsx", build_starter)]

    print(f"Building Chapter 2 workbooks from {SOURCE.name} "
          f"({len(df)} states)\n")
    for name, fn in builds:
        wb = fn(df)
        path = OUT / name
        wb.save(path)
        print(f"  {name:<32} {path.stat().st_size / 1024:6.1f} KB   "
              f"sheets: {', '.join(wb.sheetnames)}")

    g = df[VAR]
    lo, hi = float(g.min()), float(g.max())
    b = [lo + i * (hi - lo) / 6 for i in range(7)]
    dup = [m for m, v in zip(df["Member"], g) if v in [round(x, 4) for x in b[1:6]]]
    print(f"\n  The BROKEN workbook double-counts: {', '.join(dup) or 'nothing'}"
          f"  (frequencies sum to {len(g) + len(dup)}, not {len(g)})")
    print(f"  r(gPOP, HPI level)  = {g.corr(df['HPI']):+.3f}")
    print(f"  r(gPOP, gHPI)       = {g.corr(df['gHPI']):+.3f}")
    print(f"\nWritten to {OUT}")


if __name__ == "__main__":
    main()
